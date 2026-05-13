#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import json
import re
import time
import unicodedata
import urllib.parse
import urllib.request
from io import BytesIO
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - dependency guard
    load_workbook = None

try:
    import curl_cffi
except ImportError:
    pass


@dataclass(frozen=True)
class LiveMarket:
    line: float
    over: float
    under: float
    provider_id: int | None = None


@dataclass(frozen=True)
class LiveSnapshot:
    match_url: str
    event_id: int
    home_team: str
    away_team: str
    tournament: str
    tournament_slug: str
    category_name: str
    country_name: str
    status_text: str
    minute: float
    goals_home: float
    goals_away: float
    yellows_home: float
    yellows_away: float
    yellows_total: float
    reds_home: float
    reds_away: float
    reds_total: float
    fouls_home: float
    fouls_away: float
    fouls_total: float
    corners_home: float
    corners_away: float
    corners_total: float
    crosses_home: float
    crosses_away: float
    referee: str | None
    xg_home: float
    xg_away: float
    shots_home: float
    shots_away: float
    shots_on_target_home: float
    shots_on_target_away: float
    possession_home: float
    possession_away: float
    centros_local: float = 0.0
    centros_visitante: float = 0.0
    urgency_multiplier: float = 1.0
    defensive_yellows: float = 0.0
    touches_in_box_home: float = 0.0
    touches_in_box_away: float = 0.0
    dangerous_attacks_home: float = 0.0
    dangerous_attacks_away: float = 0.0
    big_chances_missed_home: float = 0.0
    big_chances_missed_away: float = 0.0
    # Attack map real de SofaScore (12 zonas 4x3, normalizadas 0-1). Lista vacía = usar estimación JS.
    attack_zones_home: tuple[float, ...] = ()
    attack_zones_away: tuple[float, ...] = ()
    goals_market: LiveMarket | None = None
    corners_market: LiveMarket | None = None
    cards_market: LiveMarket | None = None
    notes: tuple[str, ...] = ()
    home_win_odds: float | None = None
    away_win_odds: float | None = None


@dataclass(frozen=True)
class PrematchContext:
    source: str
    home_team: str
    away_team: str
    goal_total: float | None = None
    corner_total: float | None = None
    card_total: float | None = None
    goal_signal: float = 0.0
    corner_signal: float = 0.0
    card_signal: float = 0.0
    notes: tuple[str, ...] = ()


def normalize_name(value: Any) -> str:
    text = "" if value is None else str(value)
    # Reparar mojibake (Latin-1 → UTF-8) antes de normalizar
    try:
        text = text.encode('latin1').decode('utf-8')
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass
    # Quitar acentos via NFKD + strip non-ASCII
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower().replace("&", " and ")
    # Eliminar prefijos/sufijos de club que SofaScore agrega
    text = re.sub(r'\b(fc|afc|cf|ca|cd|nk|sk|us|fk|sc|fs|ac|as|rc|rcd|sd|ud|ad|sl|bk)\b', '', text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def ascii_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    text = re.sub(r"[^0-9.\-]+", "", text)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def clamp(value: float, lower: float, upper: float) -> float:
    return max(lower, min(upper, value))


def signal_from_flag(value: Any) -> float:
    text = ascii_text(value).strip().lower()
    if not text:
        return 0.0
    if "si" in text or "yes" in text or "true" in text or "✅" in str(value):
        return 1.0
    if "no" in text or "false" in text or "❌" in str(value):
        return -1.0
    return 0.0


def fractional_to_decimal(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().upper()
    if not text:
        return None
    if "/" in text:
        num_text, den_text = text.split("/", 1)
        num = safe_float(num_text)
        den = safe_float(den_text)
        if num is None or den in {None, 0.0}:
            return None
        return round((num / den) + 1.0, 2)
    direct = safe_float(text)
    if direct is not None and direct > 1.0:
        return round(direct, 2)
    if text in {"EVS", "EVENS", "EVEN"}:
        return 2.0
    return None


def best_match_score(target_home: str, target_away: str, row_home: str, row_away: str) -> int:
    score = 0
    if target_home == row_home:
        score += 60
    elif target_home in row_home or row_home in target_home:
        score += 30

    if target_away == row_away:
        score += 60
    elif target_away in row_away or row_away in target_away:
        score += 30

    return score


class PrematchWorkbook:
    def __init__(self, path: str | Path):
        if load_workbook is None:
            raise RuntimeError("Falta openpyxl. Instala openpyxl para leer el Excel prematch.")
        self.source = str(path)
        self.remote_refresh_seconds = 60.0
        self.last_refresh_monotonic = 0.0
        self.path = Path(path) if not self._is_remote_source(self.source) else None
        self.source_label = self._build_source_label(self.source)
        self._workbook_stream: BytesIO | None = None
        self.workbook = None
        self.partidos_rows: list[dict[str, Any]] = []
        self.datos_rows: list[dict[str, Any]] = []
        self.racha_goles: dict[str, dict[str, Any]] = {}
        self.racha_corners: dict[str, dict[str, Any]] = {}
        self.racha_tarjetas: dict[str, dict[str, Any]] = {}
        self._reload_data()

    def close(self) -> None:
        if self.workbook is not None:
            self.workbook.close()
        if self._workbook_stream is not None:
            self._workbook_stream.close()

    def _is_remote_source(self, value: str) -> bool:
        lowered = value.strip().lower()
        return lowered.startswith("http://") or lowered.startswith("https://")

    def _build_source_label(self, value: str) -> str:
        if not self._is_remote_source(value):
            return f"xlsx:{Path(value).name}"
        if "docs.google.com" in value and "/spreadsheets/" in value:
            return "gsheets:live"
        parsed = urllib.parse.urlparse(value)
        host = parsed.netloc or "remote"
        return f"url:{host}"

    def _google_sheets_export_url(self, value: str) -> str:
        parsed = urllib.parse.urlparse(value)
        path = parsed.path

        published_match = re.search(r"/spreadsheets/d/e/([a-zA-Z0-9-_]+)", path)
        if published_match:
            published_id = published_match.group(1)
            return f"https://docs.google.com/spreadsheets/d/e/{published_id}/pub?output=xlsx"

        standard_match = re.search(r"/spreadsheets/d/(?!e/)([a-zA-Z0-9-_]+)", path)
        if standard_match:
            sheet_id = standard_match.group(1)
            return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

        return value

    def _open_workbook_source(self, value: str) -> str | Path | BytesIO:
        if not self._is_remote_source(value):
            return Path(value)

        source_url = value
        if "docs.google.com" in value and "/spreadsheets/" in value:
            source_url = self._google_sheets_export_url(value)

        request = urllib.request.Request(source_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(request, timeout=25) as response:
            payload = response.read()
        self._workbook_stream = BytesIO(payload)
        return self._workbook_stream

    def _reload_data(self) -> None:
        if self.workbook is not None:
            self.workbook.close()
        if self._workbook_stream is not None:
            self._workbook_stream.close()
            self._workbook_stream = None

        workbook_input = self._open_workbook_source(self.source)
        self.workbook = load_workbook(workbook_input, read_only=True, data_only=True)
        self.partidos_rows = self._load_sheet_rows("PartidosHoy")
        self.datos_rows = self._load_sheet_rows("DatosHoy")
        self.racha_goles = self._build_team_index("RachaGolesHoy")
        self.racha_corners = self._build_team_index("RachaCornersHoy")
        self.racha_tarjetas = self._build_team_index("RachaTarjetasHoy")
        self.last_refresh_monotonic = time.monotonic()

    def refresh_if_needed(self) -> None:
        if not self._is_remote_source(self.source):
            return
        if (time.monotonic() - self.last_refresh_monotonic) < self.remote_refresh_seconds:
            return
        self._reload_data()

    def _load_sheet_rows(self, name: str) -> list[dict[str, Any]]:
        if name not in self.workbook.sheetnames:
            return []
        ws = self.workbook[name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = ["" if cell is None else str(cell).strip() for cell in next(rows)]
        except StopIteration:
            return []
        output: list[dict[str, Any]] = []
        for row in rows:
            if not row or not any(cell not in {None, ""} for cell in row):
                continue
            output.append({headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))})
        return output

    def _build_team_index(self, name: str) -> dict[str, dict[str, Any]]:
        rows = self._load_sheet_rows(name)
        index: dict[str, dict[str, Any]] = {}
        for row in rows:
            team_key = normalize_name(row.get("Equipo"))
            if team_key and team_key not in index:
                index[team_key] = row
        return index

    def _find_match_row(
        self,
        rows: list[dict[str, Any]],
        home_field: str,
        away_field: str,
        home_team: str,
        away_team: str,
    ) -> dict[str, Any] | None:
        home_key = normalize_name(home_team)
        away_key = normalize_name(away_team)
        best_row: dict[str, Any] | None = None
        best_score = 0
        for row in rows:
            row_home = normalize_name(row.get(home_field))
            row_away = normalize_name(row.get(away_field))
            score = best_match_score(home_key, away_key, row_home, row_away)
            if score > best_score:
                best_row = row
                best_score = score
        return best_row if best_score >= 60 else None

    def lookup(self, home_team: str, away_team: str) -> PrematchContext | None:
        self.refresh_if_needed()
        partidos_row = self._find_match_row(
            self.partidos_rows,
            "Equipo Local",
            "Equipo Visitante",
            home_team,
            away_team,
        )
        datos_row = self._find_match_row(
            self.datos_rows,
            "🏠 Local",
            "🆚 Visitante",
            home_team,
            away_team,
        )

        goal_total = None
        corner_total = None
        card_total = None
        goal_signal = 0.0
        corner_signal = 0.0
        card_signal = 0.0
        notes: list[str] = []

        if partidos_row:
            prompt = str(partidos_row.get("🔬 PROMPT CLINICO") or partidos_row.get("⚡ PROMPT EXPRESS") or "")
            parsed = self._parse_prompt_text(prompt)
            goal_total = parsed.get("goal_total")
            corner_total = parsed.get("corner_total")
            if parsed.get("goal_signal") is not None:
                goal_signal += float(parsed["goal_signal"])
            if parsed.get("corner_signal") is not None:
                corner_signal += float(parsed["corner_signal"])
            notes.extend(parsed.get("notes", []))

        if datos_row:
            goal_signal += signal_from_flag(datos_row.get("⚽ 2.5")) * 0.30
            corner_signal += signal_from_flag(datos_row.get("🚩 8.5")) * 0.25
            card_signal += signal_from_flag(datos_row.get("🟨 3.5")) * 0.25
            notes.append("Excel DatosHoy encontro senales de mercado del dia.")

        home_goal_row = self.racha_goles.get(normalize_name(home_team))
        away_goal_row = self.racha_goles.get(normalize_name(away_team))
        if goal_total is None:
            goal_vals = [safe_float(home_goal_row.get("Promedio")) if home_goal_row else None]
            goal_vals.append(safe_float(away_goal_row.get("Promedio")) if away_goal_row else None)
            goal_candidates = [value for value in goal_vals if value is not None]
            if goal_candidates:
                goal_total = sum(goal_candidates) / len(goal_candidates)
                notes.append("RachaGolesHoy aporto una expectativa de goles base.")

        home_corner_row = self.racha_corners.get(normalize_name(home_team))
        away_corner_row = self.racha_corners.get(normalize_name(away_team))
        if corner_total is None:
            corner_vals = [safe_float(home_corner_row.get("Promedio")) if home_corner_row else None]
            corner_vals.append(safe_float(away_corner_row.get("Promedio")) if away_corner_row else None)
            corner_candidates = [value for value in corner_vals if value is not None]
            if corner_candidates:
                corner_total = sum(corner_candidates) / len(corner_candidates)
                notes.append("RachaCornersHoy aporto una expectativa de corners base.")

        home_card_row = self.racha_tarjetas.get(normalize_name(home_team))
        away_card_row = self.racha_tarjetas.get(normalize_name(away_team))
        card_vals = [safe_float(home_card_row.get("Promedio")) if home_card_row else None]
        card_vals.append(safe_float(away_card_row.get("Promedio")) if away_card_row else None)
        card_candidates = [value for value in card_vals if value is not None]
        if card_candidates:
            card_total = sum(card_candidates) / len(card_candidates)
            notes.append("RachaTarjetasHoy aporto una expectativa de tarjetas base.")

        if (
            goal_total is None
            and corner_total is None
            and card_total is None
            and goal_signal == 0.0
            and corner_signal == 0.0
            and card_signal == 0.0
        ):
            return None

        ho = aw = None
        if partidos_row:
            for key in (
                "Cuota 1", "Cuota Local", "ML Local", "1", "Home ML", "Win Local",
            ):
                if ho is None:
                    v = fractional_to_decimal(partidos_row.get(key))
                    if v is not None and v > 1.0:
                        ho = v
            for key in (
                "Cuota 2", "Cuota Visitante", "ML Visitante", "2", "Away ML", "Win Visitante",
            ):
                if aw is None:
                    v = fractional_to_decimal(partidos_row.get(key))
                    if v is not None and v > 1.0:
                        aw = v

        return PrematchContext(
            source=self.source_label,
            home_team=home_team,
            away_team=away_team,
            goal_total=goal_total,
            corner_total=corner_total,
            card_total=card_total,
            goal_signal=clamp(goal_signal, -1.0, 1.0),
            corner_signal=clamp(corner_signal, -1.0, 1.0),
            card_signal=clamp(card_signal, -1.0, 1.0),
            notes=tuple(dict.fromkeys(note for note in notes if note)),
            home_win_odds=ho,
            away_win_odds=aw,
        )

    def _parse_prompt_text(self, prompt: str) -> dict[str, Any]:
        normalized = ascii_text(prompt)
        output: dict[str, Any] = {"notes": []}

        goal_match = re.search(
            r"Goles esperados\s*[-:> ]+\s*Local:\s*([0-9.]+)\s*\|\s*Visitante:\s*([0-9.]+)\s*\|\s*Total:\s*([0-9.]+)",
            normalized,
            flags=re.IGNORECASE,
        )
        if goal_match:
            output["goal_total"] = safe_float(goal_match.group(3))
            output["notes"].append("PartidosHoy trajo expectativa total de goles.")

        corner_match = re.search(r"Corners esperados:\s*([0-9.]+)", normalized, flags=re.IGNORECASE)
        if corner_match:
            output["corner_total"] = safe_float(corner_match.group(1))
            output["notes"].append("PartidosHoy trajo expectativa total de corners.")

        over25 = re.search(r"Over 2\.5 Goles:\s*([0-9]{1,3})%", normalized, flags=re.IGNORECASE)
        if over25:
            probability = safe_float(over25.group(1))
            if probability is not None:
                output["goal_signal"] = (probability - 50.0) / 50.0

        over95 = re.search(r"Over 9\.5 Corners:\s*([0-9]{1,3})%", normalized, flags=re.IGNORECASE)
        if over95:
            probability = safe_float(over95.group(1))
            if probability is not None:
                output["corner_signal"] = (probability - 50.0) / 50.0

        return output


def _search_json_item(data: Any, home_team: str | None, away_team: str | None) -> dict[str, Any] | None:
    if isinstance(data, dict):
        return data
    if not isinstance(data, list):
        return None
    if home_team is None or away_team is None:
        return next((item for item in data if isinstance(item, dict)), None)

    home_key = normalize_name(home_team)
    away_key = normalize_name(away_team)
    best_item: dict[str, Any] | None = None
    best_score = 0
    for item in data:
        if not isinstance(item, dict):
            continue
        item_home = normalize_name(
            item.get("home_team")
            or item.get("home")
            or item.get("local")
            or item.get("equipo_local")
            or (item.get("teams") or {}).get("home")
        )
        item_away = normalize_name(
            item.get("away_team")
            or item.get("away")
            or item.get("visitante")
            or item.get("equipo_visitante")
            or (item.get("teams") or {}).get("away")
        )
        score = best_match_score(home_key, away_key, item_home, item_away)
        if score > best_score:
            best_item = item
            best_score = score
    return best_item if best_score >= 60 else None


def load_prematch_context_from_json(
    source: str,
    home_team: str | None = None,
    away_team: str | None = None,
    token: str | None = None,
) -> PrematchContext | None:
    if source.startswith("http://") or source.startswith("https://"):
        headers = {"User-Agent": "Mozilla/5.0"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        request = urllib.request.Request(source, headers=headers)
        with urllib.request.urlopen(request, timeout=25) as response:
            raw = response.read().decode("utf-8")
    else:
        raw = Path(source).read_text(encoding="utf-8")

    data = json.loads(raw)
    item = _search_json_item(data, home_team, away_team)
    if not item:
        return None

    goal_total = safe_float(
        item.get("goal_total")
        or item.get("expected_goals_total")
        or item.get("goals_total")
        or item.get("xg_total")
    )
    corner_total = safe_float(
        item.get("corner_total")
        or item.get("expected_corners_total")
        or item.get("corners_total")
    )
    card_total = safe_float(
        item.get("card_total")
        or item.get("expected_cards_total")
        or item.get("cards_total")
        or item.get("tarjetas_total")
    )
    goal_signal = safe_float(item.get("goal_signal")) or 0.0
    corner_signal = safe_float(item.get("corner_signal")) or 0.0
    card_signal = safe_float(item.get("card_signal")) or 0.0
    notes = item.get("notes") or item.get("comentarios") or []
    if isinstance(notes, str):
        notes = [notes]

    if (
        goal_total is None
        and corner_total is None
        and card_total is None
        and goal_signal == 0.0
        and corner_signal == 0.0
        and card_signal == 0.0
    ):
        return None

    ho = safe_float(item.get("home_win_odds") or item.get("odds_home") or item.get("ml_home"))
    aw = safe_float(item.get("away_win_odds") or item.get("odds_away") or item.get("ml_away"))
    if ho is not None and ho <= 1.0:
        ho = None
    if aw is not None and aw <= 1.0:
        aw = None

    return PrematchContext(
        source=f"json:{source}",
        home_team=str(item.get("home_team") or item.get("home") or home_team or ""),
        away_team=str(item.get("away_team") or item.get("away") or away_team or ""),
        goal_total=goal_total,
        corner_total=corner_total,
        card_total=card_total,
        goal_signal=clamp(goal_signal, -1.0, 1.0),
        corner_signal=clamp(corner_signal, -1.0, 1.0),
        card_signal=clamp(card_signal, -1.0, 1.0),
        notes=tuple(str(note) for note in notes if note),
        home_win_odds=ho,
        away_win_odds=aw,
    )


def merge_prematch_contexts(*contexts: PrematchContext | None) -> PrematchContext | None:
    valid = [context for context in contexts if context is not None]
    if not valid:
        return None
    if len(valid) == 1:
        return valid[0]

    goal_totals = [context.goal_total for context in valid if context.goal_total is not None]
    corner_totals = [context.corner_total for context in valid if context.corner_total is not None]
    card_totals = [context.card_total for context in valid if context.card_total is not None]
    notes: list[str] = []
    for context in valid:
        notes.extend(context.notes)

    ho = aw = None
    for context in valid:
        if ho is None and getattr(context, "home_win_odds", None):
            ho = context.home_win_odds
        if aw is None and getattr(context, "away_win_odds", None):
            aw = context.away_win_odds

    return PrematchContext(
        source=" + ".join(context.source for context in valid),
        home_team=valid[0].home_team,
        away_team=valid[0].away_team,
        goal_total=(sum(goal_totals) / len(goal_totals)) if goal_totals else None,
        corner_total=(sum(corner_totals) / len(corner_totals)) if corner_totals else None,
        card_total=(sum(card_totals) / len(card_totals)) if card_totals else None,
        goal_signal=clamp(sum(context.goal_signal for context in valid) / len(valid), -1.0, 1.0),
        corner_signal=clamp(sum(context.corner_signal for context in valid) / len(valid), -1.0, 1.0),
        card_signal=clamp(sum(context.card_signal for context in valid) / len(valid), -1.0, 1.0),
        notes=tuple(dict.fromkeys(note for note in notes if note)),
        home_win_odds=ho,
        away_win_odds=aw,
    )


class SofaScoreMonitor:
    def __init__(
        self,
        provider_id: int = 1,
        headless: bool = True,  # Maintained for backwards compatibility in init signatures
        timeout_ms: int = 25000,
    ):
        self.provider_id = int(provider_id)
        self.timeout_sec = timeout_ms / 1000.0
        self._standings_cache = {}
        self._headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'es-CO,es;q=0.9,en-US;q=0.8,en;q=0.7',
            'Cache-Control': 'no-cache',
            'Origin': 'https://www.sofascore.com',
            'Sec-Ch-Ua': '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin'
        }

    def open(self) -> None:
        pass

    def close(self) -> None:
        pass

    def fetch_snapshot(self, match_url: str, reload_page: bool = True) -> LiveSnapshot:
        import re
        from curl_cffi import requests

        # Extract event ID from match_url
        # Example: https://www.sofascore.com/en-us/football/match/nice-olympique-de-marseille/QHslI#id:14167983
        event_id = None
        m = re.search(r'#id:(\d+)', match_url)
        if m:
            event_id = int(m.group(1))
        else:
            m2 = re.search(r'/(\d+)/?$', match_url)
            if m2:
                event_id = int(m2.group(1))
                
        if not event_id:
            raise ValueError(f"No se pudo extraer event_id de la URL: {match_url}")

        # Usar SIEMPRE un Referer de SofaScore, nunca de sitios externos (Betano, etc.)
        # Un Referer externo hace que Cloudflare rechace las peticiones a api.sofascore.com con 403.
        self._headers['Referer'] = f'https://www.sofascore.com/event/{event_id}'
        self._headers['Origin'] = 'https://www.sofascore.com'

        payload = {
            "event": {"status": 0, "body": {}},
            "incidents": {"status": 0, "body": {}},
            "statistics": {"status": 0, "body": {}},
            "graph": {"status": 0, "body": {}},
            "odds": {"status": 0, "body": {}},
            "attackMap": {"status": 0, "body": {}},
            "standings": {"status": 0, "body": {}},
        }

        try:
            def _fetch(endpoint: str) -> dict:
                try:
                    r = requests.get(f"https://api.sofascore.com{endpoint}", headers=self._headers, impersonate="chrome124", timeout=self.timeout_sec)
                    if r.status_code == 200:
                        r.encoding = 'utf-8'
                        return {"status": 200, "body": r.json()}
                    return {"status": r.status_code, "body": {}}
                except Exception:
                    return {"status": 0, "body": {}}

            payload["event"] = _fetch(f"/api/v1/event/{event_id}")
            if payload["event"]["status"] != 200:
                raise ValueError(f"SofaScore API retornó {payload['event']['status']} al buscar el evento {event_id}. Posible bloqueo de Cloudflare o ID inválido.")
            
            payload["incidents"] = _fetch(f"/api/v1/event/{event_id}/incidents")
            payload["statistics"] = _fetch(f"/api/v1/event/{event_id}/statistics")
            payload["graph"] = _fetch(f"/api/v1/event/{event_id}/graph")
            payload["odds"] = _fetch(f"/api/v1/event/{event_id}/odds/{self.provider_id}/all")
            payload["attackMap"] = _fetch(f"/api/v1/event/{event_id}/attack-map")  # 404 en muchos partidos
            payload["averagePositions"] = _fetch(f"/api/v1/event/{event_id}/average-positions")

            # Check if standings are needed
            initial_event = payload["event"]["body"].get("event", {})
            tournament_id = (initial_event.get("tournament") or {}).get("id")
            season_id = (initial_event.get("season") or {}).get("id")
            need_standings = False
            
            if tournament_id and season_id and (tournament_id, season_id) not in self._standings_cache:
                payload["standings"] = _fetch(f"/api/v1/tournament/{tournament_id}/season/{season_id}/standings/total")
                if payload["standings"]["status"] == 200:
                    self._standings_cache[(tournament_id, season_id)] = payload["standings"]["body"]
        except Exception as e:
            if "SofaScore API" not in str(e):
                print(f"[SOFASCORE] Fetch error: {e}")
            raise e

        event_payload = (
            payload["event"]["body"]["event"]
            if payload["event"]["status"] == 200 and payload["event"]["body"] and "event" in payload["event"]["body"]
            else initial_event
        )
        incidents_payload = payload["incidents"]["body"] if payload["incidents"]["status"] == 200 else {}
        statistics_payload = payload["statistics"]["body"] if payload["statistics"]["status"] == 200 else {}
        odds_payload = payload["odds"]["body"] if payload["odds"]["status"] == 200 else {}
        graph_payload = payload.get("graph", {}).get("body", {}) if payload.get("graph", {}).get("status") == 200 else {}

        # ── Attack Map: SofaScore attack-map (12 zonas) o Average Positions (fallback) ──
        # Coordenadas XY en escala 0-100. El campo cenital es:
        #   X: 0 = portería propia, 100 = portería rival (eje longitudinal)
        #   Y: 0 = banda izquierda, 100 = banda derecha
        # Nuestro layout 4 cols × 3 filas:
        #   col 0 = X 0-25 (def propio), col 1 = X 25-50, col 2 = X 50-75, col 3 = X 75-100 (área rival)
        #   row 0 = Y 0-33 (banda izq), row 1 = Y 33-67 (centro), row 2 = Y 67-100 (banda der)
        attack_zones_home: list[float] = []
        attack_zones_away: list[float] = []

        def _zones_from_avg_positions(players: list) -> list[float]:
            """Convierte posiciones promedio (X, Y, pointsCount) en 12 zonas normalizadas."""
            grid = [0.0] * 12
            for p in players:
                ax = float(p.get("averageX") or 0)
                ay = float(p.get("averageY") or 0)
                pts = float(p.get("pointsCount") or 1)
                col = min(3, int(ax / 25))
                row = min(2, int(ay / 33.34))
                zone_idx = row * 4 + col
                grid[zone_idx] += pts
            max_v = max(grid) or 1.0
            return [round(v / max_v, 3) for v in grid]

        # Prioridad 1: attack-map oficial (solo disponible en algunos partidos durante el juego)
        attack_map_payload = payload.get("attackMap", {})
        if attack_map_payload and attack_map_payload.get("status") == 200:
            am_body = attack_map_payload.get("body") or {}
            def _parse_zones(side_data: dict) -> list[float]:
                zones_raw = {z["zone"]: z.get("touches", 0) for z in side_data.get("zones", [])}
                vals = [float(zones_raw.get(i, 0)) for i in range(12)]
                max_v = max(vals) or 1.0
                return [round(v / max_v, 3) for v in vals]
            if am_body.get("home"):
                attack_zones_home = _parse_zones(am_body["home"])
            if am_body.get("away"):
                attack_zones_away = _parse_zones(am_body["away"])

        # Prioridad 2: average-positions (disponible en casi todos los partidos)
        if not attack_zones_home or not attack_zones_away:
            avg_pos_payload = payload.get("averagePositions", {})
            if avg_pos_payload and avg_pos_payload.get("status") == 200:
                ap_body = avg_pos_payload.get("body") or {}
                home_players = ap_body.get("home") or []
                away_players = ap_body.get("away") or []
                if home_players and not attack_zones_home:
                    attack_zones_home = _zones_from_avg_positions(home_players)
                if away_players and not attack_zones_away:
                    # Para el equipo visitante, SofaScore reporta X desde su propia portería (espejo)
                    # Invertimos X para que ambos equipos ataquen "hacia la derecha" en nuestra vista
                    mirrored = [{**p, "averageX": 100.0 - float(p.get("averageX") or 0)} for p in away_players]
                    attack_zones_away = _zones_from_avg_positions(mirrored)


        stats_map = _flatten_statistics(statistics_payload)
        # We DO NOT read yellows from stats_map because it includes manager/bench cards!
        # yellow_home=... / yellow_away=... will be calculated from incidents below.
        foul_home = _stat_pair(stats_map, "fouls")[0] or 0.0
        foul_away = _stat_pair(stats_map, "fouls")[1] or 0.0
        corner_home = _stat_pair(stats_map, "cornerKicks")[0] or 0.0
        corner_away = _stat_pair(stats_map, "cornerKicks")[1] or 0.0
        xg_home = _stat_pair(stats_map, "expectedGoals")[0] or 0.0
        xg_away = _stat_pair(stats_map, "expectedGoals")[1] or 0.0
        # xG fallback: si SofaScore no provee xG (ej: Liga Argentina), estimar desde tiros
        # Se completará DESPUÉS de parsear shots_on_target y shots_home
        shots_home = _stat_pair(stats_map, "totalShotsOnGoal")[0] or 0.0
        shots_away = _stat_pair(stats_map, "totalShotsOnGoal")[1] or 0.0
        shots_on_target_home = _stat_pair(stats_map, "shotsOnGoal")[0] or 0.0
        shots_on_target_away = _stat_pair(stats_map, "shotsOnGoal")[1] or 0.0
        possession_home = _stat_pair(stats_map, "ballPossession")[0] or 50.0
        
        # ── Punto 1 (Fix): Mapeo Universal de TIB ──────────────────────────────
        # SofaScore alterna entre 'touchesInOppBox' y 'touchesInPenaltyArea'.
        # Usamos _first_nonzero_stat para no detenernos en 0.0 (falso negativo).
        crosses_home = _first_nonzero_stat(stats_map, "crosses", "accurateCross", "totalCrosses", side=0)
        crosses_away = _first_nonzero_stat(stats_map, "crosses", "accurateCross", "totalCrosses", side=1)

        touches_in_box_home = _first_nonzero_stat(
            stats_map, "touches_in_opposition_box", "touchesInOppBox", "touchesInPenaltyArea", side=0
        )
        touches_in_box_away = _first_nonzero_stat(
            stats_map, "touches_in_opposition_box", "touchesInOppBox", "touchesInPenaltyArea", side=1
        )

        # ── Punto 2 (Fix): Proxy de Ataques Peligrosos con peso de posesión ───────
        # Si no hay 'dangerousAttacks' (Brasileirão, DIMAYOR, etc.),
        # usamos finalThirdEntries * (posesion/100) como proxy de presión real.
        possession_away = 100.0 - possession_home if possession_home > 0 else 50.0
        _da_home_raw = _first_nonzero_stat(stats_map, "dangerous_attacks", "dangerousAttacks", side=0)
        _da_away_raw = _first_nonzero_stat(stats_map, "dangerous_attacks", "dangerousAttacks", side=1)
        _fte_home = _first_nonzero_stat(stats_map, "finalThirdEntries", side=0)
        _fte_away = _first_nonzero_stat(stats_map, "finalThirdEntries", side=1)
        if _da_home_raw > 0 or _da_away_raw > 0:
            # Liga con datos nativos de ataques peligrosos
            dangerous_attacks_home = _da_home_raw
            dangerous_attacks_away = _da_away_raw
        elif _fte_home > 0 or _fte_away > 0:
            # Proxy: finalThirdEntries ponderado por posesión
            dangerous_attacks_home = round(_fte_home * (possession_home / 100.0), 1)
            dangerous_attacks_away = round(_fte_away * (possession_away / 100.0), 1)
        else:
            dangerous_attacks_home = 0.0
            dangerous_attacks_away = 0.0

        big_chances_missed_home = _first_nonzero_stat(stats_map, "bigChancesMissed", "bigChanceMissed", side=0)
        big_chances_missed_away = _first_nonzero_stat(stats_map, "bigChancesMissed", "bigChanceMissed", side=1)

        # xG fallback: si SofaScore no provee expectedGoals (ej: Liga Argentina, DIMAYOR),
        # estimarlo desde tiros con pesos basados en ubicación.
        # Formula: shots_on_target * 0.09 + shots_inside_box * 0.04 + big_chances * 0.35
        if xg_home == 0.0 and (shots_home > 0 or shots_on_target_home > 0):
            shots_inside_home = _stat_pair(stats_map, "totalShotsInsideBox")[0] or 0.0
            big_chances_home = _stat_pair(stats_map, "bigChanceCreated")[0] or 0.0
            xg_home = round(shots_on_target_home * 0.09 + shots_inside_home * 0.04 + big_chances_home * 0.35, 2)
        if xg_away == 0.0 and (shots_away > 0 or shots_on_target_away > 0):
            shots_inside_away = _stat_pair(stats_map, "totalShotsInsideBox")[1] or 0.0
            big_chances_away = _stat_pair(stats_map, "bigChanceCreated")[1] or 0.0
            xg_away = round(shots_on_target_away * 0.09 + shots_inside_away * 0.04 + big_chances_away * 0.35, 2)

        incidents = incidents_payload.get("incidents") or []
        current_minute = _compute_live_minute(event_payload, incidents)

        # Fallback for Missing Tactical Data eliminado.
        # Ahora se prioriza el uso de dangerousAttacks y ballPossession como fuente primaria de presión
        # en lugar de adivinar el TIB desde el gráfico de momentum.
        
        referee = None
        if event_payload.get("referee"):
            referee = event_payload["referee"].get("name")
        
        # Filtro estricto: Eliminar tarjetas a Managers, Banquillo o Anuladas por VAR
        def is_valid_card(inc):
            if inc.get("incidentType") != "card": return False
            if inc.get("rescinded", False): return False
            if "manager" in inc: return False
            if inc.get("isBench", False): return False
            if inc.get("time", 0) <= 0: return False  # Tiempos negativos = fuera de campo
            return True

        valid_reds = [i for i in incidents if is_valid_card(i) and str(i.get("incidentClass") or "").lower() in {"red", "yellowred", "secondyellowred"}]
        reds_home = float(sum(1 for i in valid_reds if i.get("isHome") is True))
        reds_away = float(sum(1 for i in valid_reds if i.get("isHome") is False))
        reds_total = float(reds_home + reds_away)

        valid_yellows = [i for i in incidents if is_valid_card(i) and str(i.get("incidentClass") or "").lower() == "yellow"]
        yellow_home = float(sum(1 for i in valid_yellows if i.get("isHome") is True))
        yellow_away = float(sum(1 for i in valid_yellows if i.get("isHome") is False))

        # FASE 1: Amonestados Defensivos
        yellow_incidents = valid_yellows
        
        defensive_yellows_count = 0.0
        for yinc in yellow_incidents:
            player = yinc.get("player", {})
            pos = str(player.get("position", "")).upper()
            if pos in {"D", "M"}:  # SofaScore reporta D (Defender) y M (Midfielder)
                defensive_yellows_count += 1.0

        # FASE 1: Multiplicador de Urgencia por Tabla
        urgency_multiplier = 1.0
        if tournament_id and season_id and (tournament_id, season_id) in self._standings_cache:
            try:
                standings_data = self._standings_cache[(tournament_id, season_id)]
                if standings_data and "standings" in standings_data and len(standings_data["standings"]) > 0:
                    rows = standings_data["standings"][0].get("rows", [])
                    total_rows = len(rows)
                    if total_rows > 0:
                        home_pos = next((i for i, r in enumerate(rows) if r.get("team", {}).get("id") == event_payload.get("homeTeam", {}).get("id")), -1)
                        away_pos = next((i for i, r in enumerate(rows) if r.get("team", {}).get("id") == event_payload.get("awayTeam", {}).get("id")), -1)
                        
                        # Si cualquiera de los equipos está en los últimos 3 lugares (descenso) o primeros 3 lugares (campeonato)
                        is_home_urgent = (home_pos >= total_rows - 4) or (home_pos <= 3) if home_pos != -1 else False
                        is_away_urgent = (away_pos >= total_rows - 4) or (away_pos <= 3) if away_pos != -1 else False
                        
                        if is_home_urgent or is_away_urgent:
                            urgency_multiplier = 1.25 # Factor de desesperacion
            except:
                pass

        current_goals = float((event_payload.get("homeScore") or {}).get("current") or 0.0) + float((event_payload.get("awayScore") or {}).get("current") or 0.0)
        current_corners = float(corner_home + corner_away)
        current_cards = float(yellow_home + yellow_away + reds_total)

        notes: list[str] = []
        if payload["statistics"]["status"] != 200:
            notes.append("SofaScore no devolvio estadisticas completas para este partido.")
        if payload["odds"]["status"] != 200:
            notes.append("SofaScore no devolvio cuotas live para este partido.")

        return LiveSnapshot(
            match_url=match_url,
            event_id=int(event_payload["id"]),
            home_team=str(event_payload["homeTeam"]["name"]),
            away_team=str(event_payload["awayTeam"]["name"]),
            tournament=str(event_payload["tournament"]["name"]),
            tournament_slug=str((event_payload.get("tournament") or {}).get("slug") or ""),
            category_name=str((((event_payload.get("tournament") or {}).get("category") or {}).get("name")) or ""),
            country_name=str((((event_payload.get("tournament") or {}).get("category") or {}).get("country") or {}).get("name") or ""),
            status_text=str(event_payload["status"]["description"]),
            minute=current_minute,
            goals_home=float((event_payload.get("homeScore") or {}).get("current") or 0.0),
            goals_away=float((event_payload.get("awayScore") or {}).get("current") or 0.0),
            yellows_home=float(yellow_home),
            yellows_away=float(yellow_away),
            yellows_total=float(yellow_home + yellow_away),
            reds_home=reds_home,
            reds_away=reds_away,
            reds_total=reds_total,
            fouls_home=float(foul_home),
            fouls_away=float(foul_away),
            fouls_total=float(foul_home + foul_away),
            corners_home=float(corner_home),
            corners_away=float(corner_away),
            corners_total=float(corner_home + corner_away),
            crosses_home=float(crosses_home),
            crosses_away=float(crosses_away),
            referee=referee,
            xg_home=float(xg_home),
            xg_away=float(xg_away),
            shots_home=float(shots_home),
            shots_away=float(shots_away),
            shots_on_target_home=float(shots_on_target_home),
            shots_on_target_away=float(shots_on_target_away),
            possession_home=float(possession_home),
            possession_away=100.0 - possession_home if possession_home > 0 else 50.0,
            centros_local=float(crosses_home),
            centros_visitante=float(crosses_away),
            urgency_multiplier=float(urgency_multiplier),
            defensive_yellows=float(defensive_yellows_count),
            touches_in_box_home=float(touches_in_box_home),
            touches_in_box_away=float(touches_in_box_away),
            dangerous_attacks_home=float(dangerous_attacks_home),
            dangerous_attacks_away=float(dangerous_attacks_away),
            big_chances_missed_home=float(big_chances_missed_home),
            big_chances_missed_away=float(big_chances_missed_away),
            attack_zones_home=tuple(attack_zones_home),
            attack_zones_away=tuple(attack_zones_away),
            goals_market=_extract_total_market(odds_payload, "Match goals", self.provider_id, current_goals, str(event_payload.get("tournament", {}).get("name", ""))),
            corners_market=_extract_total_market(odds_payload, "Corners 2-Way", self.provider_id, current_corners, str(event_payload.get("tournament", {}).get("name", ""))),
            cards_market=_extract_total_market(odds_payload, "Cards in match", self.provider_id, current_cards, str(event_payload.get("tournament", {}).get("name", ""))),
            notes=tuple(notes),
        )

    def fetch_event_info(self, event_id: int) -> dict:
        from curl_cffi import requests
        try:
            r = requests.get(f"https://api.sofascore.com/api/v1/event/{event_id}", headers=self._headers, impersonate="chrome124", timeout=self.timeout_sec)
            if r.status_code == 200:
                r.encoding = 'utf-8'
                return r.json()
            return {"error": str(r.status_code)}
        except Exception as e:
            return {"error": str(e)}


def _flatten_statistics(payload: dict[str, Any]) -> dict[str, tuple[float | None, float | None]]:
    output: dict[str, tuple[float | None, float | None]] = {}
    for bucket in payload.get("statistics") or []:
        if bucket.get("period") != "ALL":
            continue
        for group in bucket.get("groups") or []:
            for item in group.get("statisticsItems") or []:
                key = str(item.get("key") or item.get("name") or "").strip()
                output[key] = (safe_float(item.get("homeValue")), safe_float(item.get("awayValue")))
    return output


def _stat_pair(
    stats_map: dict[str, tuple[float | None, float | None]],
    key: str,
) -> tuple[float | None, float | None]:
    return stats_map.get(key, (None, None))


def _first_nonzero_stat(stats_map, *keys, side: int = 0) -> float:
    """
    Intenta leer las claves en orden y devuelve el PRIMER valor > 0.
    A diferencia del chain `or`, no se detiene en 0.0 (que Python evalua como falsy).
    side=0 para home, side=1 para away.
    """
    for key in keys:
        val = stats_map.get(key, (None, None))[side]
        if val is not None and val > 0:
            return float(val)
    return 0.0

def _compute_live_minute(event: dict[str, Any], incidents: list[dict[str, Any]]) -> float:
    status = event.get("status") or {}
    status_type = str(status.get("type") or "").lower()
    status_desc = str(status.get("description") or "").lower()
    time_info = event.get("time") or {}

    if status_type == "finished":
        max_seconds = safe_float(time_info.get("max")) or 5400.0
        extra_seconds = safe_float(time_info.get("extra")) or 0.0
        return round((max_seconds + extra_seconds) / 60.0, 1)

    if status_desc == "halftime":
        return 45.0

    current_period_start = safe_float(time_info.get("currentPeriodStartTimestamp"))
    initial_seconds = safe_float(time_info.get("initial")) or 0.0
    max_seconds = safe_float(time_info.get("max")) or 5400.0
    extra_seconds = safe_float(time_info.get("extra")) or 0.0

    if current_period_start is not None and status_type == "inprogress":
        elapsed_seconds = max(0.0, time.time() - current_period_start)
        minute = (initial_seconds + elapsed_seconds) / 60.0
        return round(clamp(minute, 0.0, (max_seconds + extra_seconds) / 60.0), 1)

    incident_minutes = []
    for incident in incidents:
        base = safe_float(incident.get("time"))
        added = safe_float(incident.get("addedTime")) or 0.0
        if base is not None and base >= 0:
            incident_minutes.append(base + added)
    if incident_minutes:
        return float(max(incident_minutes))

    return round((initial_seconds or 0.0) / 60.0, 1)


# Alias de nombres de mercado de SofaScore — cubre variaciones por liga/región
_MARKET_ALIASES: dict[str, tuple[str, ...]] = {
    "Match goals":   ("Match goals", "Total Goals", "Over/Under", "Goals"),
    "Corners 2-Way": ("Corners 2-Way", "Total Corners", "Corner Kicks", "Corners Over/Under"),
    "Cards in match": (
        "Cards in match", "Total Cards", "Booking Points", "Yellow Cards",
        "Total Bookings", "Cards", "Tarjetas", "Bookings",
    ),
}


def _extract_total_market(
    payload: dict[str, Any],
    market_name: str,
    provider_id: int,
    current_total: float = 0.0,
    _debug_league: str = "",
) -> LiveMarket | None:
    """
    Extrae el mercado total activo del payload de SofaScore más cercano al valor
    justo (menor imbalance over/under).

    Soporta alias de nombres de mercado para cubrir variaciones por liga/región.
    Emite un log diagnóstico cuando el mercado no se encuentra, para identificar
    qué nombres usa SofaScore y poder actualizar los alias.
    """
    # Resolver alias: intentar cada nombre candidato en orden
    names_to_try = _MARKET_ALIASES.get(market_name, (market_name,))

    candidates: list[tuple[float, float, float, float, float]] = []
    for market in payload.get("markets") or []:
        mkt_name = market.get("marketName", "")
        if mkt_name not in names_to_try:
            continue
        if not market.get("isLive") or market.get("suspended"):
            continue
        line = safe_float(market.get("choiceGroup"))
        if line is None or line <= current_total:
            continue
        over = None
        under = None
        for choice in market.get("choices") or []:
            decimal = fractional_to_decimal(
                choice.get("fractionalValue") or choice.get("initialFractionalValue")
            )
            if decimal is None:
                continue
            if choice.get("name") == "Over":
                over = decimal
            elif choice.get("name") == "Under":
                under = decimal
        if over is None or under is None:
            continue
        vig = (1.0 / over) + (1.0 / under)
        imbalance = abs(over - under)
        candidates.append((imbalance, abs(vig - 1.08), line, over, under))

    if not candidates:
        # Log diagnóstico: qué mercados live SÍ existen en este payload
        live_names = sorted({
            m.get("marketName", "")
            for m in (payload.get("markets") or [])
            if m.get("isLive") and not m.get("suspended")
        })
        if live_names:
            print(
                f"[MARKET-DIAG] '{market_name}' no encontrado"
                + (f" ({_debug_league})" if _debug_league else "")
                + f". Mercados live disponibles: {live_names[:12]}"
            )
        return None

    # ── Selección inteligente de línea ──────────────────────────────────
    # Prioriza la línea más cercana al total actual (más el margen mínimo de 0.5).
    # Esto evita que el scraper seleccione líneas absurdas (ej: OVER 9.5 con 9 corners)
    # simplemente porque tiene el menor spread.
    # Dentro de las líneas "cercanas" (diferencia < 3.0 del total actual),
    # se prefiere la de menor imbalance. Si ninguna está cerca, se usa la de menor spread.
    close_candidates = [c for c in candidates if (c[2] - current_total) <= 3.0]
    pool = close_candidates if close_candidates else candidates

    _, _, line, over, under = min(pool, key=lambda row: (row[0], row[1], row[2]))
    return LiveMarket(line=float(line), over=float(over), under=float(under), provider_id=provider_id)



